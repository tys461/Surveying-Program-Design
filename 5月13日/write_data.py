import prossces
from open import Open_dat
from dini03_parser import Dini03Parser
from openpyxl import load_workbook


def write():
    opener=Open_dat()
    raw_data=opener.list_data
    parser=Dini03Parser(raw_data)
    line_info=parser.parse()


    if line_info.PartList:
        for line in line_info.PartList:
            '''建立xlsx文件'''
            prossces.duplicate_template('data/观测手簿模版.xlsx',
                                        f'output{line_info.PartList.index(line)+1}.xlsx',
                                        "A7:J9",10,len(line.StationList)-1)

            '''打开建好的xlsx文件'''
            wb = load_workbook(f'output{line_info.PartList.index(line)+1}.xlsx')
            ws = wb.active

            part=line
            part.StarPtH=0.0
            part.Reset()


            BH=7
            FH=8
            DH=9
            R_H1 = 7
            R_H2 = 8
            print(line.StartPtName[4:6],line.EndPtName[4:6])
            print(line.StationCount)
            ws[f"C{6+line.StationCount*3+1}"]=line.StartPtName[4:6]#起始点
            ws[f"C{6+line.StationCount*3+2}"]=line.EndPtName[4:6]#结束点

            ws[f"C{6+line.StationCount*3+3}"]=f'{line.Df/1000:.5f}'#累计前距
            ws[f"C{6+line.StationCount*3+4}"]=f'{line.Db/1000:.5f}'#累计后距

            ws[f"F{6+line.StationCount*3+3}"]=f'{line.dz:.5f}'#累计高差
            ws[f"F{6+line.StationCount*3+4}"]=f'{line.dz:.5f}'#测段距离



            for station in line.StationList:
                ws[f'I{BH}']=f'{station.BPtH:.5f}'
                ws[f'I{FH}']=f'{station.FPtH:.5f}'
                ws[f'H{DH}']=f'{station.DeltH:.5f}'

                FH += 3
                BH += 3
                DH += 3


                R_L_D1=5
                R_L_D2=5
                R_L_H1=3
                R_L_H2=3
                for sight in station.SightList:
                    # print(sight.ptName,sight.RD,sight.SType)
                    if sight.SType=="B":
                        ws.cell(row=R_H1, column=R_L_D1).value =f'{sight.RD:.5f}'
                        ws.cell(row=R_H1, column=R_L_H1).value = f'{sight.HD:.5f}'
                        R_L_D1+=1
                        R_L_H1+=1
                    if sight.SType == "F":
                        ws.cell(row=R_H2, column=R_L_D2).value = f'{sight.RD:.5f}'
                        ws.cell(row=R_H2, column=R_L_H2).value = f'{sight.HD:.5f}'
                        R_L_D2+=1
                        R_L_H2+=1
                R_H1+=3
                R_H2+=3

            wb.save(f'output{line_info.PartList.index(line)+1}.xlsx')

write()