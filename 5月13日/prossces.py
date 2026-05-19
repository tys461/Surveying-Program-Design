from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy

def copy_range_down(ws, source_range, start_row, copies=1, copy_style=True):
    """
    将包含合并单元格的区域向下复制多份（自动下移下方内容）

    :param ws: 工作表对象
    :param source_range: 源区域，如 "A7:J9"
    :param start_row: 第一份副本的起始行（必须 > 源区域最大行）
    :param copies: 复制份数
    :param copy_style: 是否复制样式
    :return: None
    """
    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    src_height = max_row - min_row + 1
    src_width = max_col - min_col + 1

    if start_row <= max_row:
        raise ValueError(f"起始行 {start_row} 必须大于源区域最大行 {max_row}")

    # 收集源区域内的合并单元格（相对偏移）
    src_merges = []
    for merged in ws.merged_cells.ranges:
        if (merged.min_row >= min_row and merged.max_row <= max_row and
            merged.min_col >= min_col and merged.max_col <= max_col):
            src_merges.append((
                merged.min_row - min_row, merged.max_row - min_row,
                merged.min_col - min_col, merged.max_col - min_col
            ))

    # 收集起始行下方的合并单元格，先删除，插入行后再恢复
    affected_merges = [merged for merged in ws.merged_cells.ranges if merged.min_row >= start_row]
    for merged in affected_merges:
        ws.unmerge_cells(str(merged))

    insert_rows = src_height * copies
    ws.insert_rows(start_row, amount=insert_rows)

    # 恢复受影响的合并单元格（整体下移）
    for merged in affected_merges:
        new_min_row = merged.min_row + insert_rows
        new_max_row = merged.max_row + insert_rows
        ws.merge_cells(start_row=new_min_row, start_column=merged.min_col,
                       end_row=new_max_row, end_column=merged.max_col)

    # 确定需要复制的单元格（排除合并区域中非左上角的单元格）
    cells_to_copy = set()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            # 检查当前单元格是否属于某个合并区域且不是左上角
            is_non_top_left = False
            for rel_min_r, rel_max_r, rel_min_c, rel_max_c in src_merges:
                abs_min_r = min_row + rel_min_r
                abs_max_r = min_row + rel_max_r
                abs_min_c = min_col + rel_min_c
                abs_max_c = min_col + rel_max_c
                if abs_min_r <= r <= abs_max_r and abs_min_c <= c <= abs_max_c:
                    if (r, c) != (abs_min_r, abs_min_c):
                        is_non_top_left = True
                    break
            if not is_non_top_left:
                cells_to_copy.add((r, c))

    # 复制多份
    for copy_idx in range(copies):
        target_start_row = start_row + copy_idx * src_height
        for src_r, src_c in cells_to_copy:
            src_cell = ws.cell(row=src_r, column=src_c)
            tgt_r = target_start_row + (src_r - min_row)
            tgt_cell = ws.cell(row=tgt_r, column=src_c)
            tgt_cell.value = src_cell.value
            if copy_style and src_cell.has_style:
                # 更完整的样式复制
                if src_cell.font:
                    tgt_cell.font = copy(src_cell.font)
                if src_cell.fill:
                    tgt_cell.fill = copy(src_cell.fill)
                if src_cell.border:
                    tgt_cell.border = copy(src_cell.border)
                if src_cell.alignment:
                    tgt_cell.alignment = copy(src_cell.alignment)
                if src_cell.number_format:
                    tgt_cell.number_format = src_cell.number_format
                # 可选：protection
                if src_cell.protection:
                    tgt_cell.protection = copy(src_cell.protection)

        # 创建本副本的合并单元格
        for rel_min_r, rel_max_r, rel_min_c, rel_max_c in src_merges:
            new_min_row = target_start_row + rel_min_r
            new_max_row = target_start_row + rel_max_r
            new_min_col = min_col + rel_min_c
            new_max_col = min_col + rel_max_c
            if new_min_row == new_max_row and new_min_col == new_max_col:
                continue
            ws.merge_cells(start_row=new_min_row, start_column=new_min_col,
                           end_row=new_max_row, end_column=new_max_col)


def duplicate_template(template_path, output_path, source_range, start_row, copies, copy_style=True):
    """高层封装：加载模板 → 复制区域 → 保存"""
    wb = load_workbook(template_path)
    ws = wb.active
    copy_range_down(ws, source_range, start_row, copies, copy_style)
    wb.save(output_path)


# 使用示例
