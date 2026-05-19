import re
from openpyxl import load_workbook
from openpyxl.styles import numbers
from typing import Optional, List
from typing import List

class SightInfo:
    """单次观测信息类"""
    def __init__(self, line_id: int = 0, pt_name: str = "", rd: float = 0.0, hd: float = 0.0,
                 adr: int = 0, time_str: str = "", s_type: str = "") -> None:
        self.LineID = line_id      # 所在原始数据行号
        self.ptName = pt_name
        self.RD = rd
        self.HD = hd
        self.Adr = adr
        self.TimeStr = time_str
        self.SType = s_type        # 测量类型：后视B，前视F


class StationInfo:
    """测站信息类"""
    def __init__(self) -> None:
        self.FPtName = ""          # 前视点名
        self.FPtH = 0.0            # 前视点高程
        self.BPtName = ""          # 后视点名
        self.BPtH = 0.0            # 后视点高程
        self.SightList: List[SightInfo] = []   # 测量信息
        self.Rb1 = self.Rb2 = self.Rf1 = self.Rf2 = 0.0  # 后/前视中丝读数
        self.Db1 = self.Db2 = self.Df1 = self.Df2 = 0.0  # 后/前视距
        self.DeltH = 0.0           # 测站高差
        self.DeltD = 0.0           # 测站视距差

    def Reset(self) -> None:
        """重新设置测站信息（后视点高程BPtH需在调用前赋值）"""
        r_num = f_num = 0

        for sight in self.SightList:
            if sight.SType == "B":          # 后视
                r_num += 1
                if r_num == 1:              # 第1次
                    self.Rb1 = sight.RD
                    self.Db1 = sight.HD
                else:                       # 第2次
                    self.Rb2 = sight.RD
                    self.Db2 = sight.HD
                self.BPtName = sight.ptName
            else:                           # 前视
                f_num += 1
                if f_num == 1:
                    self.Rf1 = sight.RD
                    self.Df1 = sight.HD
                else:
                    self.Rf2 = sight.RD
                    self.Df2 = sight.HD
                self.FPtName = sight.ptName

        # 测站高差和视距差（观测次数成对，SightList长度应为偶数）
        self.DeltH = (self.Rb1 - self.Rf1 + self.Rb2 - self.Rf2) / (len(self.SightList) / 2)
        self.DeltD = (self.Db1 - self.Df1 + self.Db2 - self.Df2) / (len(self.SightList) / 2)

        # 计算前视点高程
        self.FPtH = self.DeltH + self.BPtH


class PartInfo:
    """测段信息类"""
    def __init__(self) -> None:
        self.partName = ""                   # 测段名
        self.partNum = 0                     # 测段号
        self.PartCode = ""                   # 测段代码
        self.StartPtName = ""                # 开始点名
        self.EndPtName = ""                  # 结束点名
        self.StationList: List[StationInfo] = []   # 测站信息
        self.StationCount = 0                # 测站数
        self.StartPtH = 0.0                  # 开始点高程
        self.EndPtH = 0.0                    # 结束点高程
        self.sh = 0.0                        # 累计高差
        self.dz = 0.0                        # 线路闭合差
        self.Db = 0.0                        # 累计后视距
        self.Df = 0.0                        # 累计前视距

    def Reset(self) -> None:
        """重新设置测段信息（第一个测站的后视点高程需事先赋值为StartPtH）"""
        self.partName = f"{self.StartPtName}-{self.EndPtName}"
        self.StationCount = len(self.StationList)
        self.sh = self.Db = self.Df = 0.0

        for i, station in enumerate(self.StationList):
            if i > 0:
                station.BPtH = self.StationList[i - 1].FPtH
            station.Reset()
            self.sh += station.DeltH
            # 累计后/前视距：每个测站的后/前视距取平均值（观测次数为2次时，分母=2）
            self.Db += (station.Db1 + station.Db2) / (len(station.SightList) / 2)
            self.Df += (station.Df1 + station.Df2) / (len(station.SightList) / 2)

        self.dz = self.EndPtH - (self.StartPtH + self.sh)


class LineInfo:
    """线路信息类，对应一个观测文件中的所有测段信息"""
    def __init__(self) -> None:
        self.LineName = ""                  # 线路名
        self.PartList: List[PartInfo] = []  # 测段信息

    def GetLength(self) -> float:
        """获得线路长度（后视距+前视距之和）"""
        length = 0.0
        for part in self.PartList:
            length += part.Db + part.Df
        return length

# ================== 实体类 ==================
# 使用之前转换的 Python 类（略，请将前文定义的 SightInfo, StationInfo, PartInfo, LineInfo 复制在此）
# 确保类中已定义 __init__ 和 Reset 方法

# ================== DAT 文件解析 ==================
def parse_dini03_file(filepath: str) -> LineInfo:
    """
    解析天宝 DINI03 的 .dat 文本文件，返回 LineInfo 对象。
    注意：返回后需要手动设置每个 PartInfo 的 StartPtH 和 EndPtH，
         然后调用 PartInfo.Reset() 进行高程传递和闭合差计算。
    """
    line_info = LineInfo()
    line_info.LineName = filepath

    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    in_section = False          # 是否在测段内
    current_part: Optional[PartInfo] = None
    current_station: Optional[StationInfo] = None

    for raw_line in lines:
        line = raw_line.strip()
        if not line.startswith("For M5"):
            continue

        parts = line.split('|')
        if len(parts) < 3:
            continue

        field2 = parts[2].strip()
        field3 = parts[3].strip() if len(parts) > 3 else ""
        field4 = parts[4].strip() if len(parts) > 4 else ""
        field5 = parts[5].strip() if len(parts) > 5 else ""

        # ---------- 测段开始 ----------
        if "TO  Start-Line" in field2:
            if current_part is not None and current_part.StationList:
                line_info.PartList.append(current_part)
            current_part = PartInfo()
            # 提取测段号
            match = re.search(r'\b(\d+)\b', field2)
            if match:
                current_part.partNum = int(match.group(1))
            # 提取测段代码
            code_match = re.search(r'[A-Za-z]+', field2)
            if code_match:
                current_part.PartCode = code_match.group()
            in_section = True
            current_station = None
            continue

        # ---------- 测段结束 ----------
        if "TO  End-Line" in field2:
            if in_section and current_part is not None:
                if current_station is not None and current_station.SightList:
                    current_part.StationList.append(current_station)
                line_info.PartList.append(current_part)
                current_part = None
            in_section = False
            continue

        # 忽略控制行
        if not in_section or current_part is None:
            continue
        if any(key in field2 for key in ["TO  Cont-Line", "TO  Reading", "TO  Station repeated"]):
            continue

        # ---------- 观测数据行 (Rb / Rf) ----------
        if field3.startswith("Rb") or field3.startswith("Rf"):
            sight = SightInfo()
            sight.SType = field3[:2]   # "Rb" 或 "Rf"
            # 读数 RD
            rd_match = re.search(r'(\d+\.\d+)', field3)
            if rd_match:
                sight.RD = float(rd_match.group(1))
            # 视距 HD
            hd_match = re.search(r'(\d+\.\d+)', field4)
            if hd_match:
                sight.HD = float(hd_match.group(1))
            # 点名和时间
            tokens = field2.split()
            if len(tokens) >= 2:
                pt = tokens[1].split('#')[0]   # 去除 '#'
                sight.ptName = pt
            if len(tokens) >= 3:
                sight.TimeStr = tokens[2]
            # 行号作为 LineID
            adr_match = re.search(r'Adr\s+(\d+)', parts[1])
            if adr_match:
                sight.LineID = int(adr_match.group(1))

            if current_station is None:
                current_station = StationInfo()
            current_station.SightList.append(sight)
            continue

        # ---------- 测站结束标志 (Z 行，非 Sh/dz) ----------
        if field5.startswith("Z") and not field3.startswith("Sh") and not field3.startswith("dz"):
            if current_station is not None and current_station.SightList:
                current_part.StationList.append(current_station)
                current_station = None
            continue

        # 其他汇总行忽略（Sh, dz, Db, Df 等会在 Reset 时重新计算）

    # 文件结束处理
    if current_part is not None and current_part.StationList:
        if current_station is not None and current_station.SightList:
            current_part.StationList.append(current_station)
        line_info.PartList.append(current_part)

    return line_info


# ================== Excel 写入 ==================
def write_part_to_excel(template_path: str, output_path: str, part: PartInfo, start_h: float, end_h: float):
    """
    将单个测段的数据写入 Excel 模板。
    :param template_path: 模板文件路径
    :param output_path:  输出文件路径
    :param part:         测段对象（已设置 StartPtName, EndPtName 等，但尚未调用 Reset）
    :param start_h:      测段起始点已知高程 (m)
    :param end_h:        测段结束点已知高程 (m)
    """
    # 设置高程并执行计算
    part.StartPtH = start_h
    part.EndPtH = end_h
    part.Reset()   # 计算各测站高差、高程等

    wb = load_workbook(template_path)
    ws = wb.active

    # --- 1. 查找数据起始行（定位 "后视" 单元格）---
    data_start_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == "后视":   # B列通常为"视准点"下的内容
            data_start_row = row
            break
    if data_start_row is None:
        raise ValueError("模板中未找到'后视'行，请检查模板结构")

    # 每个测站占用的行数：后视行、前视行、视距差/累积差行，之后加一个空行（可选）
    rows_per_station = 3
    current_row = data_start_row

    # 累积视距差（单位: m）
    cumulative_dist_diff = 0.0

    # --- 2. 写入各测站数据 ---
    for idx, station in enumerate(part.StationList, start=1):
        # 后视行
        ws.cell(current_row, 1, idx)                     # 测站号
        ws.cell(current_row, 2, "后视")
        ws.cell(current_row, 3, station.Db1)             # 后距1 (m)
        ws.cell(current_row, 4, station.Db2)             # 后距2 (m)
        ws.cell(current_row, 5, station.Rb1)             # 后尺读数1 (m)
        ws.cell(current_row, 6, station.Rb2)             # 后尺读数2 (m)
        # 读数差 (mm) = |Rb1 - Rb2| * 1000
        ws.cell(current_row, 7, abs(station.Rb1 - station.Rb2) * 1000)
        ws.cell(current_row, 8, "")                      # 高差列留空
        ws.cell(current_row, 9, station.BPtH)            # 后视点高程 (已知)
        ws.cell(current_row, 10, "")                     # 备注

        # 前视行（下一行）
        ws.cell(current_row + 1, 1, idx)
        ws.cell(current_row + 1, 2, "前视")
        ws.cell(current_row + 1, 3, station.Df1)         # 前距1 (m)
        ws.cell(current_row + 1, 4, station.Df2)         # 前距2 (m)
        ws.cell(current_row + 1, 5, station.Rf1)         # 前尺读数1 (m)
        ws.cell(current_row + 1, 6, station.Rf2)         # 前尺读数2 (m)
        ws.cell(current_row + 1, 7, abs(station.Rf1 - station.Rf2) * 1000)  # 读数差
        ws.cell(current_row + 1, 8, station.DeltH)       # 本测站高差
        ws.cell(current_row + 1, 9, station.FPtH)        # 前视点高程 (计算值)
        ws.cell(current_row + 1, 10, "")

        # 计算视距差 (m) 和累积视距差 (m)
        dist_diff = station.DeltD   # 测站视距差
        cumulative_dist_diff += dist_diff

        # 视距差/累积差行（再下一行）
        ws.cell(current_row + 2, 1, idx)
        ws.cell(current_row + 2, 2, "")
        ws.cell(current_row + 2, 3, dist_diff)           # 视距差(m)
        ws.cell(current_row + 2, 4, cumulative_dist_diff)  # 累积差(m)
        # 第5-6列填写两次高差（同 station.DeltH）?
        ws.cell(current_row + 2, 5, station.DeltH)       # 高差(m) 第一次
        ws.cell(current_row + 2, 6, station.DeltH)       # 高差(m) 第二次
        ws.cell(current_row + 2, 7, "")                  # 读数差留空
        ws.cell(current_row + 2, 8, "")                  # 高差列留空
        ws.cell(current_row + 2, 9, "")                  # 高程留空
        ws.cell(current_row + 2, 10, "")

        # 移动行指针到下一组（跳过一行空行以增加可读性）
        current_row += rows_per_station + 1

    # --- 3. 写入测段汇总信息 ---
    # 查找“测段”区域起始行（通常有"测段起点"文字）
    part_start_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value and "测段起点" in str(ws.cell(row, 1).value):
            part_start_row = row
            break
    if part_start_row is None:
        # 若找不到，则附加在所有数据之后
        part_start_row = current_row + 2
        ws.cell(part_start_row, 1, "测段起点")
        ws.cell(part_start_row + 1, 1, "测段终点")
        ws.cell(part_start_row + 2, 1, "累计前距")
        ws.cell(part_start_row + 3, 1, "累计后距")
        ws.cell(part_start_row + 4, 1, "累计高差")
        ws.cell(part_start_row + 5, 1, "测段距离")
    else:
        # 清理原有内容，避免残留
        for i in range(6):
            for col in range(1, 10):
                ws.cell(part_start_row + i, col, None)

    ws.cell(part_start_row, 2, part.StartPtName)
    ws.cell(part_start_row + 1, 2, part.EndPtName)
    ws.cell(part_start_row + 2, 2, part.Df / 1000)          # 累计前距 (km)
    ws.cell(part_start_row + 2, 3, "km")
    ws.cell(part_start_row + 3, 2, part.Db / 1000)          # 累计后距 (km)
    ws.cell(part_start_row + 3, 3, "km")
    ws.cell(part_start_row + 4, 2, part.sh)                 # 累计高差 (m)
    ws.cell(part_start_row + 4, 3, "m")
    ws.cell(part_start_row + 5, 2, (part.Db + part.Df) / 1000)  # 测段距离 (km)
    ws.cell(part_start_row + 5, 3, "km")

    # 可选：写入“累计视距差”到指定位置（模板中可能有专门单元格）
    # 此处暂时忽略，用户可自行扩展。

    # 保存文件
    wb.save(output_path)
    print(f"已生成手簿: {output_path}")


# ================== 使用示例 ==================
if __name__ == "__main__":
    # 1. 解析 .dat 文件
    dat_file = "0624CH67#连续梁.DAT"
    line = parse_dini03_file(dat_file)

    # 2. 输出共有几个测段
    print(f"共解析出 {len(line.PartList)} 个测段")

    # 3. 假设第一个测段的已知起始/结束高程（示例值，实际需根据控制点填写）
    if line.PartList:
        part0 = line.PartList[0]
        part0.StartPtName = "X1"      # 根据文件点名，可能需要从数据中提取
        part0.EndPtName = "X3"
        # 假设已知起点高程 100.000 m，终点高程 100.000 m（闭合环）
        start_h = 100.000
        end_h = 100.000 + part0.sh   # 若闭合，sh 应为 0，则 end_h = start_h
        # 实际终点高程应根据已知水准点设置，这里仅作示例
        write_part_to_excel("观测手簿模版2.xlsx", "output_测段1.xlsx", part0, start_h, end_h)